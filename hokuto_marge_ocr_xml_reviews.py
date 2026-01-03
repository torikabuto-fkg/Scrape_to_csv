#!/usr/bin/env python3
"""
OCRデータとXML抽出データをマージして高精度なレビュー抽出を行う

OCRデータ: 「良い点」「気になる点」の見出しを含む
XMLデータ: 本文が正確だが見出しが欠落している

両方をマッチングして、XMLの本文にOCRの見出しを追加する
"""

import json
import re
import sys
from matplotlib import lines
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from difflib import SequenceMatcher

# ===== ノイズ除去の設定（恣意的にON/OFFできる）=====
TRUNCATE_AFTER_REPORT_LINE = True   # 「ロコミ/口コミの問題を報告」を見つけたら以降を捨てる
REMOVE_MD_HEADINGS = True           # ## 良い点 / ## 気になる点 を消す

REPORT_LINE_RE = re.compile(r'(ロコミ|口コミ)の問題を報告')

# 「この行は本文ではない」扱いで行ごと消すパターン
NOISE_LINE_RES = [
    re.compile(r'^\s*もっと見る\s*$'),
    re.compile(r'^\s*全ての口コミを読む.*$'),
    re.compile(r'^\s*口コミの問題を報告.*$'),
    re.compile(r'^\s*ロコミの問題を報告.*$'),

    # 画像/ページマーカー系（行だけ消す：ここで本文を切らない）
    re.compile(r'^\s*=+\s*.*\s*=+\s*$'),
    re.compile(r'.*review_page_.*\.png.*', re.IGNORECASE),
    re.compile(r'.*\.png.*', re.IGNORECASE),

    # 区切り線
    re.compile(r'^\s*-{3,}\s*$'),

    # いわゆる見出しだけの行（本文ではない）
    re.compile(r'^\s*#+\s*$'),
]

# 「## あり/なし」「: / ：あり」などの見出し“単独行”を全部消す
MD_HEADING_RE = re.compile(r'^\s*#{1,6}\s*(良い点|気になる点)\s*$', re.UNICODE)


def clean_review_text(text: str) -> str:
    """
    レビュー本文からノイズを除去。
    - 行単位でノイズを落とす
    - 「ロコミ/口コミの問題を報告」行が出たら、そのレビュー内の残りは捨てる（設定でON/OFF）
    - Markdown見出し（## 良い点 等）も落とす（設定でON/OFF）
    """
    if not text or not isinstance(text, str):
        return ''

    text = text.replace('\r\n', '\n').replace('\r', '\n')

    cleaned_lines = []
    for raw_line in text.split('\n'):
        line = raw_line.strip()
        if not line:
            continue

        # ここに来たら以降は捨てる（“次のレビューまで”はパーサ側で処理するが、セル内はこれでOK）
        if TRUNCATE_AFTER_REPORT_LINE and REPORT_LINE_RE.search(line):
            break

        # Markdown見出しを消す
        if REMOVE_MD_HEADINGS and MD_HEADING_RE.match(line):
            continue

        # 行ノイズ判定
        is_noise = False
        for rex in NOISE_LINE_RES:
            if rex.match(line):
                is_noise = True
                break
        if is_noise:
            continue

        cleaned_lines.append(line)

    # 連続空行の正規化（ここでは空行を入れてないので不要だが一応）
    out = '\n'.join(cleaned_lines).strip()
    return out

STRICT_UNIVERSITY_FILTER = True  # Trueなら「大学名っぽくないもの」は空欄にする

UNIV_DENY_KEYWORDS = [
    '口コミ', 'ロコミ', '問題を報告', 'もっと見る', '全ての口コミ', '病院情報', '病院・研修',
    'review_page_', '.png', '採用で', 'p)', '見学した', 'マッチした', '年度', '男性', '女性'
]

# 「大学名らしさ」の最低条件（必要なら追加してOK）
UNIV_ALLOW_RE = re.compile(r'(大学|医科大学|医科歯科大学|大学校|医学部|医大)')

# 文字列中から大学名っぽい部分だけ抜き出す（最後に出た候補を採用）
UNIV_EXTRACT_RE = re.compile(
    r'([^\s　]{2,60}?(?:医科歯科大学|医科大学|大学校|大学|医大)(?:医学部)?)'
)

def normalize_university_name(raw: str) -> str:
    """大学名以外っぽいものを除去して、大学名候補だけ返す（なければ空欄）"""
    if not raw or not isinstance(raw, str):
        return ''

    # ざっくりノイズ判定（元の文字列にノイズが混ざってたら落とす）
    if any(k in raw for k in UNIV_DENY_KEYWORDS):
        return ''

    # OCRの余計な空白を潰してから候補抽出
    compact = re.sub(r'\s+', '', raw)

    # 「大学/医大/医学部」等が含まれないなら大学名ではない扱い
    if not UNIV_ALLOW_RE.search(compact):
        return '' if STRICT_UNIVERSITY_FILTER else raw.strip()

    candidates = UNIV_EXTRACT_RE.findall(compact)
    if not candidates:
        return '' if STRICT_UNIVERSITY_FILTER else raw.strip()

    cand = candidates[-1].strip()

    # 「大学」だけみたいな弱すぎる候補は落とす
    if cand in ('大学', '医大', '医学部'):
        return ''

    # 長すぎるのもノイズ扱い（必要なら閾値調整）
    if len(cand) > 40:
        return ''

    return cand

# ===== 分裂レビューの結合設定 =====
MERGE_ADJACENT_OVERLAP = True
MIN_OVERLAP_NORM_CHARS = 40      # 重複とみなす最小一致長（空白除去後の文字数）
MAX_OVERLAP_NORM_CHARS = 300     # 探す重複長の上限（速度対策）
FUZZY_OVERLAP_RATIO = 0.93       # 完全一致しない場合の許容（OCR誤差用）
HIGH_SIM_RATIO = 0.92            # ほぼ同じ文章のときは長い方を採用


def _norm_no_ws(s: str) -> str:
    return re.sub(r'\s+', '', s or '')


def _index_after_norm_chars(original: str, norm_chars: int) -> int:
    """originalの先頭から、空白除去後の文字をnorm_chars分進めた位置(インデックス)を返す"""
    if norm_chars <= 0:
        return 0
    cnt = 0
    for i, ch in enumerate(original):
        if not ch.isspace():
            cnt += 1
            if cnt >= norm_chars:
                return i + 1
    return len(original)


def _best_overlap_len(a_norm: str, b_norm: str,
                      min_len: int, max_len: int, fuzzy_ratio: float) -> int:
    """
    a_normのsuffix と b_normのprefix の重なり長を返す（空白除去済み前提）
    完全一致が無ければ、fuzzy_ratio以上なら許容
    """
    limit = min(len(a_norm), len(b_norm), max_len)
    if limit < min_len:
        return 0

    for l in range(limit, min_len - 1, -1):
        s1 = a_norm[-l:]
        s2 = b_norm[:l]
        if s1 == s2:
            return l
        # OCR誤差を許容したい場合のみ
        if SequenceMatcher(None, s1, s2).ratio() >= fuzzy_ratio:
            return l
    return 0


def merge_text_by_overlap(a: str, b: str,
                          min_overlap: int = MIN_OVERLAP_NORM_CHARS,
                          max_overlap: int = MAX_OVERLAP_NORM_CHARS,
                          fuzzy_ratio: float = FUZZY_OVERLAP_RATIO):
    """
    a と b が重複しているなら重複部分をカットして連結した文字列を返す。
    戻り値: (merged_text, merged_bool)
    """
    a = a or ''
    b = b or ''
    a_norm = _norm_no_ws(a)
    b_norm = _norm_no_ws(b)

    if not a_norm:
        return b, True
    if not b_norm:
        return a, True

    # 片方が完全に含まれる → 長い方だけ残す
    if a_norm in b_norm:
        return b, True
    if b_norm in a_norm:
        return a, True

    # 方向1: aの末尾とbの先頭が重なる
    l1 = _best_overlap_len(a_norm, b_norm, min_overlap, max_overlap, fuzzy_ratio)
    # 方向2: bの末尾とaの先頭が重なる（順序が逆だった場合の救済）
    l2 = _best_overlap_len(b_norm, a_norm, min_overlap, max_overlap, fuzzy_ratio)

    if l1 == 0 and l2 == 0:
        return a, False

    # より大きい重なりを採用
    if l1 >= l2:
        cut = _index_after_norm_chars(b, l1)
        suffix = b[cut:].lstrip()
        joiner = '\n' if (a and not a.endswith('\n') and suffix) else ''
        return (a.rstrip() + joiner + suffix), True
    else:
        cut = _index_after_norm_chars(a, l2)
        suffix = a[cut:].lstrip()
        joiner = '\n' if (b and not b.endswith('\n') and suffix) else ''
        return (b.rstrip() + joiner + suffix), True


def _same_meta_for_adjacent_merge(r1: dict, r2: dict) -> bool:
    """隣同士をマージしてよいかの最低条件（誤結合防止）"""
    if r1.get('year') != r2.get('year'):
        return False
    if r1.get('grade') != r2.get('grade'):
        return False
    if r1.get('participation') != r2.get('participation'):
        return False

    # 性別が両方埋まっていて違うなら別人扱い
    g1 = (r1.get('gender') or '').strip()
    g2 = (r2.get('gender') or '').strip()
    if g1 and g2 and g1 != g2:
        return False

    # 大学名が両方埋まっていて違うなら別人扱い（※空欄は許容）
    u1 = (r1.get('university') or '').strip()
    u2 = (r2.get('university') or '').strip()
    if u1 and u2 and u1 != u2:
        return False

    return True


def merge_adjacent_overlapping_reviews(reviews: list[dict]) -> list[dict]:
    """
    OCR由来の「前後で分裂したレビュー」を結合する。
    - 隣の行同士だけ見る（誤結合を減らす）
    - good_points同士、concerns同士の重なりだけ結合
    - ほぼ同じ文章なら長い方を採用
    """
    if not reviews:
        return reviews

    out = []
    i = 0
    while i < len(reviews):
        cur = dict(reviews[i])

        j = i + 1
        while j < len(reviews):
            nxt = reviews[j]

            if not _same_meta_for_adjacent_merge(cur, nxt):
                break

            changed = False

            # good_points の重なり
            if (cur.get('good_points') and nxt.get('good_points')):
                merged, ok = merge_text_by_overlap(cur['good_points'], nxt['good_points'])
                if ok:
                    cur['good_points'] = merged
                    changed = True
                else:
                    # ほぼ同じなら長い方だけ残す
                    if text_similarity(cur['good_points'], nxt['good_points']) >= HIGH_SIM_RATIO:
                        if len(_norm_no_ws(nxt['good_points'])) > len(_norm_no_ws(cur['good_points'])):
                            cur['good_points'] = nxt['good_points']
                        changed = True

            # concerns の重なり
            if (cur.get('concerns') and nxt.get('concerns')):
                merged, ok = merge_text_by_overlap(cur['concerns'], nxt['concerns'])
                if ok:
                    cur['concerns'] = merged
                    changed = True
                else:
                    if text_similarity(cur['concerns'], nxt['concerns']) >= HIGH_SIM_RATIO:
                        if len(_norm_no_ws(nxt['concerns'])) > len(_norm_no_ws(cur['concerns'])):
                            cur['concerns'] = nxt['concerns']
                        changed = True

            # どちらも結合できないならここで終了（隣以外までは追わない）
            if not changed:
                break

            # メタ情報の穴埋め（空欄優先で埋める）
            if not (cur.get('university') or '').strip():
                cur['university'] = nxt.get('university', '')
            if not (cur.get('gender') or '').strip():
                cur['gender'] = nxt.get('gender', '')

            j += 1

        out.append(cur)
        i = j

    return out

EXPLICIT_HEAD_RE = re.compile(r'(?m)^\s*(?:#{1,6}\s*)?(良い点|気になる点)\s*(?:[:：]\s*)?$', re.UNICODE)

def split_by_explicit_headings(text: str):
    if not text:
        return None
    t = re.sub(r'<br\s*/?>', '\n', text, flags=re.IGNORECASE)
    hits = list(EXPLICIT_HEAD_RE.finditer(t))
    if not hits:
        return None

    # 1個だけなら、その見出し以降を全部その側へ
    if len(hits) == 1:
        h = hits[0]
        body = t[h.end():].strip()
        if h.group(1) == '良い点':
            return {'good_points': body, 'concerns': ''}
        else:
            return {'good_points': '', 'concerns': body}

    # 複数あるなら、最初の2個だけで分割（通常は良い点→気になる点）
    h1, h2 = hits[0], hits[1]
    part1 = t[h1.end():h2.start()].strip()
    part2 = t[h2.end():].strip()

    if h1.group(1) == '良い点' and h2.group(1) == '気になる点':
        return {'good_points': part1, 'concerns': part2}
    if h1.group(1) == '気になる点' and h2.group(1) == '良い点':
        return {'good_points': part2, 'concerns': part1}

    return None

def load_jsonl(file_path):
    """JSONLファイルを読み込む"""
    data = []
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            if line.strip():
                data.append(json.loads(line))
    return data


def extract_hospital_info_xml(lines):
    """XML抽出データから病院情報を抽出"""
    info = {}
    info_markers = {
        'マッチ者数/定員': r'マッチ者数/定員\((\d{4})年\)',
        '強い科': r'強い科',
        '上級医の主な出身大学': r'上級医の主な出身大学',
        '病床数': r'病床数',
        '給与': r'給与',
        '救急指定': r'救急指定',
        '病院見学情報URL': r'病院見学情報URL'
    }
    
    for i, line in enumerate(lines[:100]):
        for key, pattern in info_markers.items():
            if re.search(pattern, line):
                if key == 'マッチ者数/定員' and i + 1 < len(lines):
                    info[key] = lines[i + 1].strip()
                elif key in ['強い科', '上級医の主な出身大学'] and i + 1 < len(lines):
                    info[key] = lines[i + 1].strip()
                elif key in ['病床数', '給与', '救急指定', '病院見学情報URL'] and i + 1 < len(lines):
                    info[key] = lines[i + 1].strip()
    
    return info


def parse_ocr_reviews(ocr_data):
    """OCRデータからレビューを抽出（見出し基準で分割、良い点・気になる点を別行に）"""
    all_sections = []
    
    # 年度パターン
    year_pattern = r'(?:([^\n]+?)\s+)?(\d+年)\s+(?:(男性|女性)\s+)?(見学した|マッチした|オンライン説明会に参加した|説明会に参加した)\s+(\d{4}年度)'
    
    # まず全ページから見出しセクションを収集
    for page_idx, page_data in enumerate(ocr_data):
        text = page_data['text']
        
        # 良い点・気になる点の見出しを全て見つける（###付きにも対応）
        heading_pattern = r'(?m)^\s*(?:#{2,3}\s*)?(良い点|気になる点)\s*(?:[:：]\s*)?$'
        headings = list(re.finditer(heading_pattern, text))

        for i, heading in enumerate(headings):
            heading_type = heading.group(1)  # "良い点" or "気になる点"
            
            # この見出しの前にある年度パターンを探す（最大500文字前まで）
            search_start = max(0, heading.start() - 500)
            pre_text = text[search_start:heading.start()]
            
            # 最も近い年度パターンを見つける
            year_matches = list(re.finditer(year_pattern, pre_text))
            if not year_matches:
                continue  # 年度パターンが見つからない場合はスキップ
            
            last_year_match = year_matches[-1]  # 最も近い（最後の）マッチ
            
            university = last_year_match.group(1).strip() if last_year_match.group(1) else ''
            
            # ノイズテキストをフィルタリング
            noise_patterns = ['口コミの問題を報告', 'ロコミの問題を報告', '採用で', 'p)', '全ての口コミを読む', 'もっと見る', '病院情報', '病院・研修']
            if any(noise in university for noise in noise_patterns):
                university = ''
            
            # ★追加
            university = normalize_university_name(university)

            grade = last_year_match.group(2)
            gender = last_year_match.group(3) if last_year_match.group(3) else ''
            participation = last_year_match.group(4)
            year = last_year_match.group(5)
            
            # この見出しから次の見出しまで（または終端まで）のコンテンツを取得
            content_start = heading.end()
            if i + 1 < len(headings):
                content_end = headings[i + 1].start()
            else:
                content_end = len(text)
            
            content = text[content_start:content_end]
            
            # ノイズ除去（行単位 + 報告行以降トリム + 見出し除去）
            content = clean_review_text(content)

            # 空でないコンテンツのみ保存
            if content:
                all_sections.append({
                    'university': university,
                    'grade': grade,
                    'gender': gender,
                    'participation': participation,
                    'year': year,
                    'heading_type': heading_type,
                    'content': content,
                    'page': page_idx,
                    'pos': heading.start(),
                })
    
    # 各セクションを個別のレビューとして出力（良い点・気になる点は別行）
    reviews = []
    
    for section in all_sections:
        # 良い点のみ or 気になる点のみのレビューとして作成
        if section['heading_type'] == '良い点':
            reviews.append({
                'university': section['university'],
                'grade': section['grade'],
                'gender': section['gender'],
                'participation': section['participation'],
                'year': section['year'],
                'good_points': section['content'],
                'concerns': '',
                'source': 'OCR',
                'page': section.get('page', 0),
                'pos': section.get('pos', 0),
            })
        elif section['heading_type'] == '気になる点':
            reviews.append({
                'university': section['university'],
                'grade': section['grade'],
                'gender': section['gender'],
                'participation': section['participation'],
                'year': section['year'],
                'good_points': '',
                'concerns': section['content'],
                'source': 'OCR'
            })
    
    # ===== ここから差し替え =====

    # できればページ順に並べる（page/posを入れてないならこのsortは実質そのまま）
    reviews.sort(key=lambda r: (r.get('page', 0), r.get('pos', 0)))

    # 1) 隣同士の「重なり」を結合して断裂を直す
    if MERGE_ADJACENT_OVERLAP:
        reviews = merge_adjacent_overlapping_reviews(reviews)

    # 2) それでも残る「ほぼ同じ」重複を除去（長い方を残す）
    final_reviews = []

    def same_meta(a, b):
        if a.get('year') != b.get('year'):
            return False
        if a.get('grade') != b.get('grade'):
            return False
        if a.get('participation') != b.get('participation'):
            return False

        ga = (a.get('gender') or '').strip()
        gb = (b.get('gender') or '').strip()
        if ga and gb and ga != gb:
            return False

        ua = (a.get('university') or '').strip()
        ub = (b.get('university') or '').strip()
        if ua and ub and ua != ub:
            return False

        return True

    def combined_norm(r):
        return re.sub(r'\s+', '', (r.get('good_points', '') + '\n' + r.get('concerns', '')).strip())

    for r in reviews:
        r_norm = combined_norm(r)
        if not r_norm:
            continue

        merged = False

        # 画面またぎ想定なので「直近数件」だけ見れば十分（誤結合も減る）
        for idx in range(max(0, len(final_reviews) - 3), len(final_reviews)):
            prev = final_reviews[idx]
            if not same_meta(prev, r):
                continue

            p_norm = combined_norm(prev)

            # 片方がもう片方に含まれる → 長い方だけ残す
            if p_norm in r_norm:
                final_reviews[idx] = r
                merged = True
                break
            if r_norm in p_norm:
                merged = True
                break

            # ほぼ同じ → 長い方
            if SequenceMatcher(None, p_norm, r_norm).ratio() >= HIGH_SIM_RATIO:
                if len(r_norm) > len(p_norm):
                    final_reviews[idx] = r
                merged = True
                break

        if not merged:
            final_reviews.append(r)

    return final_reviews

    # ===== ここまで差し替え =====



def load_xml_extracted_data(file_path):
    """XML抽出データを読み込む"""
    with open(file_path, 'r', encoding='utf-8') as f:
        return [line.rstrip('\n') for line in f]

TABLE_ALIGN_RE = re.compile(r'^\s*\|\s*:?-{2,}.*\|\s*$', re.UNICODE)

def normalize_xml_line(line: str) -> str:
    if not line:
        return ''
    # <br> を改行に
    line = re.sub(r'<br\s*/?>', '\n', line, flags=re.IGNORECASE)

    # Markdown表のアライメント行は捨てる
    if TABLE_ALIGN_RE.match(line):
        return ''

    # | で始まる表行は、パイプをスペースにして平文化
    if line.lstrip().startswith('|'):
        line = line.strip().strip('|')
        line = re.sub(r'\s*\|\s*', ' ', line)

    return line.strip()

def parse_xml_reviews(lines):
    """XML抽出データからレビューを抽出（見出しなし）"""
    reviews = []

    # 病院情報セクションの終了を見つける
    start_idx = 0
    for i, raw in enumerate(lines):
        line = normalize_xml_line(raw)
        if not line:
            continue
        if re.search(r'修正・追加する|口コミを書く', line):
            start_idx = i + 1
            break

    # 年度パターン（大学名を含む可能性）
    year_pattern = re.compile(
        r'^(?:([^\s]+(?:大学|医科大学|医学部|医科歯科大学|科学大学))\s+)?'
        r'(\d+)年\s+(男性\s+|女性\s+)?'
        r'(見学した|マッチした|オンライン説明会に参加した|説明会に参加した)\s+'
        r'(\d{4})年度'
    )

    i = start_idx
    while i < len(lines):
        line = normalize_xml_line(lines[i])

        # ★ 追加：空なら次へ（iを進めないと詰まることがある）
        if not line:
            i += 1
            continue

        # 年度行を検出
        match = year_pattern.match(line)
        if not match:
            i += 1
            continue

        university = match.group(1) if match.group(1) else ''
        university = normalize_university_name(university)
        grade = match.group(2) + '年'
        gender = match.group(3).strip() if match.group(3) else ''
        participation = match.group(4)
        year = match.group(5) + '年度'

        content_lines = []

        # ★重要：同じ行に「年度 + 本文」が入っている場合、残りを本文として拾う
        rest = line[match.end():].strip()
        if rest:
            # normalize_xml_lineで <br> を \n にしているなら分割して入れる
            for part in rest.split('\n'):
                part = part.strip()
                if part:
                    content_lines.append(part)

        i += 1

        # 次の年度行または終端まで本文を収集
        while i < len(lines):
            next_line = normalize_xml_line(lines[i])

            # ★ 追加：空ならスキップ（iを進めて次へ）
            if not next_line:
                i += 1
                continue

            # 次のレビューに到達
            if year_pattern.match(next_line):
                break

            # 「ロコミ/口コミの問題を報告」を見つけたら、そのレビュー本文はここで終了
            if REPORT_LINE_RE.search(next_line):
                i += 1
                # 次の年度行（次レビュー）までスキップ
                while i < len(lines):
                    probe = normalize_xml_line(lines[i])
                    if probe and year_pattern.match(probe):
                        break
                    i += 1
                break

            # その他の軽いノイズはスキップ
            if any(x in next_line for x in [
                '全ての口コミを読む', '病院情報',
                '総合点', '学歴フィルター', '忙しさ', 'ハイポ', 'ハイパー', '研修スタイル',
                'もっと見る'
            ]):
                i += 1
                continue

            # 通常本文として追加（\n を含む可能性があるので分割して入れるのが安全）
            for part in next_line.split('\n'):
                part = part.strip()
                if part:
                    content_lines.append(part)

            i += 1

        all_text = '\n'.join(content_lines)
        all_text = clean_review_text(all_text)  # 最後に共通クリーニング

        if all_text:
            reviews.append({
                'university': university,
                'grade': grade,
                'gender': gender,
                'participation': participation,
                'year': year,
                'text': all_text,
                'source': 'XML'
            })

    return reviews



def text_similarity(text1, text2):
    """2つのテキストの類似度を計算（0-1）"""
    # 空白と改行を正規化
    text1 = re.sub(r'\s+', '', text1)
    text2 = re.sub(r'\s+', '', text2)
    
    return SequenceMatcher(None, text1, text2).ratio()


def split_xml_by_ocr_structure(xml_text, ocr_good, ocr_concern):
    """OCRの見出し構造を使ってXMLテキストを分割（片側のみの場合は全量寄せる）"""
    ocr_good_len = len(re.sub(r'\s+', '', ocr_good or ''))
    ocr_concern_len = len(re.sub(r'\s+', '', ocr_concern or ''))
    total_ocr_len = ocr_good_len + ocr_concern_len

    # OCR情報がゼロなら分割不能：全部良い点側に寄せる（従来挙動）
    if total_ocr_len == 0:
        return {'good_points': xml_text, 'concerns': ''}

    # ★重要：片側しか無いなら分割せず、その側に全量寄せる
    if ocr_good_len == 0 and ocr_concern_len > 0:
        return {'good_points': '', 'concerns': xml_text}

    if ocr_concern_len == 0 and ocr_good_len > 0:
        return {'good_points': xml_text, 'concerns': ''}

    # ここから先は「両方ある」場合だけ比率分割
    xml_len = len(re.sub(r'\s+', '', xml_text))
    good_ratio = ocr_good_len / total_ocr_len
    target_split = int(xml_len * good_ratio)

    paragraphs = [p.strip() for p in xml_text.split('\n') if p.strip()]
    if len(paragraphs) <= 1:
        # 段落が無い/1個しかない場合は比率分割できないので良い点側に寄せる
        return {'good_points': xml_text, 'concerns': ''}

    cumulative_len = 0
    best_split_idx = 1
    min_diff = float('inf')

    for i, para in enumerate(paragraphs):
        cumulative_len += len(re.sub(r'\s+', '', para))
        diff = abs(cumulative_len - target_split)
        if diff < min_diff:
            min_diff = diff
            best_split_idx = i + 1

    return {
        'good_points': '\n'.join(paragraphs[:best_split_idx]),
        'concerns': '\n'.join(paragraphs[best_split_idx:])
    }



def merge_reviews(ocr_reviews, xml_reviews):
    """OCRレビューとXMLレビューをマージ（高精度版）"""
    merged = []
    matched_xml_indices = set()
    
    print(f"\n🔗 レビューをマッチング中...")
    print(f"   OCRレビュー: {len(ocr_reviews)}件")
    print(f"   XMLレビュー: {len(xml_reviews)}件")
    
    # OCRレビューを基準にマッチング
    for ocr_review in ocr_reviews:
        best_match = None
        best_similarity = 0.0
        best_xml_idx = -1
        
        # 同じ年度・参加形態のXMLレビューを探す
        for xml_idx, xml_review in enumerate(xml_reviews):
            if xml_idx in matched_xml_indices:
                continue
            
            # メタデータが一致するか確認（大学名は類似度チェックで判断）
            if (ocr_review['grade'] == xml_review['grade'] and
                ocr_review['participation'] == xml_review['participation'] and
                ocr_review['year'] == xml_review['year']):
                
                # テキスト類似度を計算
                combined_ocr = ocr_review['good_points'] + ocr_review['concerns']
                similarity = text_similarity(combined_ocr, xml_review['text'])
                
                if similarity > best_similarity:
                    best_similarity = similarity
                    best_match = xml_review
                    best_xml_idx = xml_idx
        
        # マッチング結果を処理
        if best_match and best_similarity > 0.3:  # 30%以上の類似度
            matched_xml_indices.add(best_xml_idx)
            
            # OCRの見出し構造でXMLの高精度テキストを分割
            split_result = split_xml_by_ocr_structure(
                best_match['text'],
                ocr_review['good_points'],
                ocr_review['concerns']
            )
            
            merged.append({
                'university': ocr_review['university'],
                'grade': ocr_review['grade'],
                'gender': ocr_review['gender'] or best_match['gender'],
                'participation': ocr_review['participation'],
                'year': ocr_review['year'],
                'good_points': split_result['good_points'],
                'concerns': split_result['concerns'],
                'source': 'OCR structure + XML text',
                'similarity': f"{best_similarity:.2%}"
            })
            print(f"   ✓ マッチ: {ocr_review['year']} {ocr_review['grade']} (類似度: {best_similarity:.2%}) [XML高精度テキスト使用]")
        else:
            # マッチしない場合はOCRのみ使用
            merged.append({
                'university': ocr_review['university'],
                'grade': ocr_review['grade'],
                'gender': ocr_review['gender'],
                'participation': ocr_review['participation'],
                'year': ocr_review['year'],
                'good_points': ocr_review['good_points'],
                'concerns': ocr_review['concerns'],
                'source': 'OCR only',
                'similarity': 'N/A'
            })
            print(f"   ⚠ マッチなし: {ocr_review['year']} {ocr_review['grade']} (OCRのみ使用)")
    
    # ★ ヒューリスティック分類は品質が低いため除外
    # マッチしなかったXMLレビューは追加しない
    # for xml_idx, xml_review in enumerate(xml_reviews):
    #     if xml_idx not in matched_xml_indices:
    #         # ヒューリスティックで良い点・気になる点を推定
    #         split_result = split_by_heuristics(xml_review['text'])
    #         
    #         merged.append({
    #             'university': xml_review['university'],
    #             'grade': xml_review['grade'],
    #             'gender': xml_review['gender'],
    #             'participation': xml_review['participation'],
    #             'year': xml_review['year'],
    #             'good_points': split_result['good_points'],
    #             'concerns': split_result['concerns'],
    #             'source': 'XML (heuristic split)',
    #             'similarity': 'N/A'
    #         })
    #         print(f"   + 追加: {xml_review['year']} {xml_review['grade']} (ヒューリスティック分類)")
    
    return merged


def split_by_heuristics(text):
    explicit = split_by_explicit_headings(text)
    if explicit:
        return explicit
    """ヒューリスティックで良い点・気になる点を分類"""
    paragraphs = [p.strip() for p in text.split('\n') if p.strip()]
    
    if not paragraphs:
        return {'good_points': '', 'concerns': ''}
    
    # 戦略1: ネガティブマーカー検出
    negative_markers = [
        '一方で、', '一方、', '一方で', 'しかし、', 'しかし', 'ただし、', 'ただし',
        'しかしながら、', 'しかしながら', 'ただ、', 'ただ', 'また、',
        '気になる点としては', 'デメリットとしては', '課題も', '懸念点としては'
    ]
    
    for i, para in enumerate(paragraphs):
        for marker in negative_markers:
            if para.startswith(marker) or marker in para[:50]:
                return {
                    'good_points': '\n'.join(paragraphs[:i]),
                    'concerns': '\n'.join(paragraphs[i:])
                }
    
    # 戦略2: パラグラフ数で分割（3つ以上なら中間点）
    if len(paragraphs) >= 3:
        split_point = len(paragraphs) // 2
        return {
            'good_points': '\n'.join(paragraphs[:split_point]),
            'concerns': '\n'.join(paragraphs[split_point:])
        }
    
    # 戦略3: 長文なら中間で分割
    if len(text) > 2000:
        mid_point = len(text) // 2
        split_pos = text.rfind('\n', 0, mid_point)
        if split_pos > 0:
            return {
                'good_points': text[:split_pos].strip(),
                'concerns': text[split_pos:].strip()
            }
    
    # フォールバック: 全て良い点
    return {
        'good_points': '\n'.join(paragraphs),
        'concerns': ''
    }


def create_excel(reviews, hospital_info, output_path):
    """Excelファイルを作成"""
    
    def clean_cell_value(value):
        """Excelセル用にテキストをクリーンアップ"""
        if not value or not isinstance(value, str):
            return value
        
        # 制御文字を除去（タブ、改行、キャリッジリターンは保持）
        import unicodedata
        cleaned = ''.join(char for char in value if unicodedata.category(char)[0] != 'C' or char in '\t\n\r')
        
        # = で始まる文字列は数式と誤認されるので ' を先頭に追加
        if cleaned.startswith('='):
            cleaned = "'" + cleaned
        
        # @ で始まる文字列も数式と誤認される可能性がある
        if cleaned.startswith('@'):
            cleaned = "'" + cleaned
        
        # + や - で始まる場合も数式と誤認される可能性
        if cleaned.startswith(('+', '-')) and len(cleaned) > 1 and cleaned[1].isdigit():
            cleaned = "'" + cleaned
        
        return cleaned
    
    wb = Workbook()
    
    # レビューシート
    ws_reviews = wb.active
    ws_reviews.title = "Reviews"
    
    # ヘッダー
    headers = ['大学', '学年', '性別', '参加形態', '年度', '良い点', '気になる点', 'データソース', '類似度']
    ws_reviews.append(headers)
    
    # ヘッダーのスタイル
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    for cell in ws_reviews[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # データ行
    for review in reviews:
        good = clean_review_text(review.get('good_points', ''))
        conc = clean_review_text(review.get('concerns', ''))

        ws_reviews.append([
            clean_cell_value(normalize_university_name(review.get('university', ''))),
            clean_cell_value(review['grade']),
            clean_cell_value(review.get('gender', '')),
            clean_cell_value(review['participation']),
            clean_cell_value(review['year']),
            clean_cell_value(good),
            clean_cell_value(conc),
            clean_cell_value(review.get('source', '')),
            review.get('similarity', '')
        ])

    # 列幅調整
    ws_reviews.column_dimensions['A'].width = 25
    ws_reviews.column_dimensions['B'].width = 10
    ws_reviews.column_dimensions['C'].width = 8
    ws_reviews.column_dimensions['D'].width = 20
    ws_reviews.column_dimensions['E'].width = 12
    ws_reviews.column_dimensions['F'].width = 80
    ws_reviews.column_dimensions['G'].width = 80
    ws_reviews.column_dimensions['H'].width = 25
    ws_reviews.column_dimensions['I'].width = 12
    
    # テキスト折り返し
    for row in ws_reviews.iter_rows(min_row=2, max_row=ws_reviews.max_row, min_col=6, max_col=7):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 病院情報シート
    ws_info = wb.create_sheet("Hospital Info")
    ws_info.append(['項目', '内容'])
    ws_info['A1'].fill = header_fill
    ws_info['A1'].font = header_font
    ws_info['B1'].fill = header_fill
    ws_info['B1'].font = header_font
    
    for key, value in hospital_info.items():
        ws_info.append([clean_cell_value(key), clean_cell_value(value)])
    
    ws_info.column_dimensions['A'].width = 30
    ws_info.column_dimensions['B'].width = 100
    
    # セルの文字制限を確認して長すぎるテキストを切り詰める
    for row in ws_reviews.iter_rows(min_row=2, max_row=ws_reviews.max_row):
        for cell in row:
            if cell.value and isinstance(cell.value, str) and len(cell.value) > 32000:
                cell.value = cell.value[:32000] + '...'
    
    # 保存（エラーハンドリング付き）
    try:
        wb.save(output_path)
        print(f"✅ Excelファイルを保存しました: {output_path}")
    except Exception as e:
        print(f"⚠ Excel保存エラー: {e}")
        # 代替ファイル名で保存を試みる
        import os
        alt_path = output_path.replace('.xlsx', '_backup.xlsx')
        print(f"📝 代替パスで保存を試みます: {alt_path}")
        wb.save(alt_path)
        print(f"✅ バックアップとして保存しました: {alt_path}")


def main():
    if len(sys.argv) < 3:
        print("使用方法: python merge_ocr_xml_reviews.py <OCR_JSONL> <XML_TXT> [OUTPUT_FILE]")
        print("例: python merge_ocr_xml_reviews.py さいたま赤十字_OCR.jsonl さいたま赤十字.txt [さいたま赤十字_merged.xlsx]")
        sys.exit(1)
    
    ocr_file = sys.argv[1]
    xml_file = sys.argv[2]
    
    # 出力ファイル名を生成（第3引数があればそれを使用）
    if len(sys.argv) >= 4:
        output_file = sys.argv[3]
    else:
        base_name = re.sub(r'(_OCR)?\.jsonl$', '', ocr_file)
        base_name = re.sub(r'\.txt$', '', base_name)
        output_file = f"{base_name}_merged.xlsx"
    
    print(f"📖 OCRデータを読み込み中: {ocr_file}")
    ocr_data = load_jsonl(ocr_file)
    print(f"   {len(ocr_data)}ページ分のデータを読み込みました")
    
    print(f"\n📖 XML抽出データを読み込み中: {xml_file}")
    xml_lines = load_xml_extracted_data(xml_file)
    print(f"   {len(xml_lines)}行のデータを読み込みました")
    
    print(f"\n🏥 病院情報を抽出中...")
    hospital_info = extract_hospital_info_xml(xml_lines)
    print(f"   {len(hospital_info)}項目を抽出しました")
    
    print(f"\n📝 OCRレビューを解析中...")
    ocr_reviews = parse_ocr_reviews(ocr_data)
    print(f"   {len(ocr_reviews)}件のレビューを抽出しました")
    good_count = sum(1 for r in ocr_reviews if r['good_points'])
    concern_count = sum(1 for r in ocr_reviews if r['concerns'])
    print(f"   良い点あり: {good_count}/{len(ocr_reviews)}")
    print(f"   気になる点あり: {concern_count}/{len(ocr_reviews)}")
    
    print(f"\n📝 XMLレビューを解析中...")
    xml_reviews = parse_xml_reviews(xml_lines)
    print(f"   {len(xml_reviews)}件のレビューを抽出しました")
    
    # マージ
    merged_reviews = merge_reviews(ocr_reviews, xml_reviews)
    
    print(f"\n📊 統計:")
    print(f"   総レビュー数: {len(merged_reviews)}")
    print(f"   良い点あり: {sum(1 for r in merged_reviews if r['good_points'])}/{len(merged_reviews)}")
    print(f"   気になる点あり: {sum(1 for r in merged_reviews if r['concerns'])}/{len(merged_reviews)}")
    
    # データソース別統計
    source_counts = {}
    for review in merged_reviews:
        source = review.get('source', 'unknown')
        source_counts[source] = source_counts.get(source, 0) + 1
    
    print(f"\n📌 データソース別:")
    for source, count in sorted(source_counts.items()):
        print(f"   {source}: {count}件")
    
    print(f"\n💡 精度について:")
    print(f"   OCR structure + XML text: OCRの見出し構造 + XMLの高精度本文")
    print(f"   OCR only: 画像認識テキスト（誤字の可能性あり）")
    
    print(f"\n📊 Excelファイルを作成中...")
    create_excel(merged_reviews, hospital_info, output_file)
    
    print(f"\n✨ 完了!")


if __name__ == "__main__":
    main()
